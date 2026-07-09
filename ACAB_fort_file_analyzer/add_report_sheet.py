from openpyxl import load_workbook

wb = load_workbook('Informe131I.xlsx')
if 'Reporte_131I' in wb.sheetnames:
    wb.remove(wb['Reporte_131I'])
ws = wb.create_sheet('Reporte_131I')

ws['A1'] = 'Reporte ¹³¹I'
ws['A3'] = 'Parámetro'
ws['B3'] = 'Valor'
ws['A4'] = 'XNORM'
ws['B4'] = '=IFERROR(VALUE(TRIM(\'inp.5\'!C36)), VALUE(\'inp.5\'!C36))'
ws['A5'] = 'FLUX [n/cm²/s]'
ws['B5'] = '=IFERROR(VALUE(TRIM(\'inp.5\'!B22)), VALUE(\'inp.5\'!B22))'
ws['A6'] = 'PHI total'
ws['B6'] = '=B4*B5'
ws['A7'] = 'T_IRR_h'
ws['B7'] = '=LET(vals, TOCOL(\'inp.5\'!B30:E34,TRUE), nums, FILTER(VALUE(TRIM(vals)), vals<>""), IFERROR(INDEX(nums,1), 0))'
ws['A8'] = 'T_COOL_h'
ws['B8'] = '=LET(vals, TOCOL(\'inp.5\'!B30:E34,TRUE), nums, FILTER(VALUE(TRIM(vals)), vals<>""), IFERROR(MAX(nums), 0))'
ws['A9'] = 'I131 T½ [s]'
ws['B9'] = '=LET(line, FILTER(config!A:A, LEFT(TRIM(config!A:A), 4)="I131"), txt, TRIM(TEXTAFTER(line, ":")), num, VALUE(LEFT(txt, FIND(" ", txt)-1)), unit, TRIM(MID(txt, FIND(" ", txt)+1, 99)), SWITCH(unit, "ns", num/1e9, "µs", num/1e6, "us", num/1e6, "ms", num/1e-3, "s", num, "m", num*60, "h", num*3600, "d", num*86400, "y", num*365.25*86400, 0))'
ws['A10'] = 'I131 T½ [d]'
ws['B10'] = '=IF(B9=0, "", B9/86400)'

ws['A12'] = 'I131 raw lines from fort.6'
ws['A13'] = 'Raw text'
ws['B13'] = 'Tokens'
ws['A14'] = '=FILTER(\'fort.6\'!A:A, LEFT(TRIM(\'fort.6\'!A:A),4)="I131")'
ws['B14'] = '=TEXTSPLIT(TRIM(A14#), " ", , TRUE)'

ws['A22'] = 'Resumen de I131 en fort.6'
rows = [2877, 4138, 5405, 6666, 7933, 9200]
for i, row in enumerate(rows, start=23):
    ws[f'A{i}'] = f'I131 fila {row}'
    ws[f'B{i}'] = f'=TRIM(\'fort.6\'!A{row})'
    ws[f'C{i}'] = f'=IFERROR(INDEX(TEXTSPLIT(B{i}," ",,TRUE),2), "")'
    ws[f'D{i}'] = f'=IFERROR(INDEX(TEXTSPLIT(B{i}," ",,TRUE),3), "")'

ws['A31'] = 'Enfriamiento 1: tiempos 0/0.25..2.25 h'
ws['A32'] = 'Tiempo [h]'
ws['B32'] = '=VSTACK(0, DROP(TEXTSPLIT(TRIM(\'fort.6\'!A3087), " ",, TRUE), 2))'
ws['A33'] = 'I131 row 4138'
ws['B33'] = '=DROP(TEXTSPLIT(TRIM(\'fort.6\'!A4138), " ",, TRUE), 1)'
ws['A34'] = 'I131 row 6666'
ws['B34'] = '=DROP(TEXTSPLIT(TRIM(\'fort.6\'!A6666), " ",, TRUE), 1)'

ws['A37'] = 'Enfriamiento 2: tiempos 0/2.50..3.00 h'
ws['A38'] = 'Tiempo [h]'
ws['B38'] = '=VSTACK(0, 0, DROP(TEXTSPLIT(TRIM(\'fort.6\'!A6882), " ",, TRUE), 2))'
ws['A39'] = 'I131 row 7933'
ws['B39'] = '=DROP(TEXTSPLIT(TRIM(\'fort.6\'!A7933), " ",, TRUE), 1)'

ws['A42'] = 'Enfriamiento 3: tiempos 0/2.50..3.00 h'
ws['A43'] = 'Tiempo [h]'
ws['B43'] = '=VSTACK(0, 0, DROP(TEXTSPLIT(TRIM(\'fort.6\'!A8149), " ",, TRUE), 2))'
ws['A44'] = 'I131 row 9200'
ws['B44'] = '=DROP(TEXTSPLIT(TRIM(\'fort.6\'!A9200), " ",, TRUE), 1)'

for col, width in [('A', 30), ('B', 70), ('C', 20), ('D', 20), ('E', 20)]:
    ws.column_dimensions[col].width = width

wb.save('Informe131I.xlsx')
